from pushbullet import Pushbullet
from openpyxl import load_workbook
import datetime

current_time = datetime.datetime.now()

file = "Push_test.txt"              # File containing the slouching alert message
eod_file = "Back-Up EOD.txt"        # File containing the cumulative slouch message
with open(file, mode='r') as f:
    text = f.read()                 # Reading the first file and storing it as a string named 'text'
with open(eod_file, mode='r') as f:
    eod_text = f.read()             # Reading the second file and storing it as a string named 'eod_text'

n_ppl = 3 # The number of ppl in the frame

workbook = load_workbook(filename = "Back-Up API person data.xlsx")
file_name = "Back-Up API person data.xlsx"
sheet = workbook.active     # Selecting the active sheet

for row in sheet.iter_rows(min_row=2, min_col=6, max_row=(n_ppl+1), max_col=6):     # Traversing the rows 2 to n_ppl+1 in column no 6
    for cell in row:
        if (cell.value == "Y"): # Y is mark that the person is slouching
            API_KEY = sheet.cell(row=cell.row, column=3).value # Getting value of API Key from the sheet
            pb = Pushbullet(API_KEY)

            Hey = "Hey "
            Backup = ", Back-Up!"
            Name_call = Hey + sheet.cell(row=cell.row, column=2).value + Backup # Joining the stings

            push = pb.push_note(Name_call, text) # Pushing notification with Title 'Name_call' and Body 'text'

            sheet.cell(row=cell.row, column=4).value += 1
            sheet.cell(row=cell.row, column=5).value += 2
            # 2 is the number of minutes of slouching before alerting
            workbook.save(filename=file_name)

if (current_time.hour >= 20): # Summary assuming office closing time as 8pm
    for row in sheet.iter_rows(min_row=2, min_col=5, max_row=(n_ppl + 1), max_col=5):
        for cell in row:
            API_KEY_2 = sheet.cell(row=cell.row, column=3).value
            pb = Pushbullet(API_KEY_2)

            dupli = cell.value
            s0 = chr(int(dupli % 10) + 48)
            dupli/=10
            s1 = chr(int(dupli % 10) + 48)
            dupli/=10
            s2 = chr(int(dupli % 10) + 48)

            if(s2 == "0"):
                s2 = ""
                if(s1 == "0"):
                    s1 = ""

            slouch = ""
            seperate = [s2,s1,s0]

            slouch = slouch.join(seperate)
            push = pb.push_note(slouch, eod_text)
