from pushbullet import Pushbullet
from openpyxl import load_workbook
import datetime

def push_notify():
    current_time = datetime.datetime.now() 
    file = "Push_test.txt"              
    eod_file = "Back-Up EOD.txt"        
    with open(file, mode='r') as f:
        text = f.read()                 
    with open(eod_file, mode='r') as f:
        eod_text = f.read()             
    workbook = load_workbook(filename = "Back-Up API person data.xlsx")
    file_name = "Back-Up API person data.xlsx"
    sheet = workbook.active 

    n_ppl = sheet.cell(row=2, column=16).value 
    for row in sheet.iter_rows(min_row=2, min_col=6, max_row=(n_ppl+1), max_col=6):
        for cell in row:
            if (cell.value == "Y"):
                API_KEY = sheet.cell(row=cell.row, column=3).value
                pb = Pushbullet(API_KEY)
                Hey = "Hey "
                Backup = ", Back-Up!"
                Name_call = Hey + sheet.cell(row=cell.row, column=2).value + Backup
                push = pb.push_note(Name_call, text) 
                sheet.cell(row=cell.row, column=4).value += 1
                sheet.cell(row=cell.row, column=5).value += 2
                workbook.save(filename=file_name)

    """if (current_time.hour >= 20): # Summary assuming office closing time as 8pm
        for row in sheet.iter_rows(min_row=2, min_col=5, max_row=(n_ppl + 1), max_col=5):
            for cell in row:
                API_KEY_2 = sheet.cell(row=cell.row, column=3).value
                pb = Pushbullet(API_KEY_2)

                dupli = cell.value
                s0 = chr(int(dupli % 10) + 48)
                dupli/=10
                s1 = chr(int(dupli % 10) + 48)
                dupli/=10
                s2 = chr(int(dupli % 10) + 48)

                if(s2 == "0"):
                    s2 = ""
                    if(s1 == "0"):
                        s1 = ""

                slouch = ""
                seperate = [s2,s1,s0]

                slouch = slouch.join(seperate)
                push = pb.push_note(slouch, eod_text)"""

