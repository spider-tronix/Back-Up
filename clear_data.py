from openpyxl import load_workbook
def clear(file_name):
    workbook = load_workbook(filename = file_name)
    sheet = workbook.active
    clear_num = sheet.cell(row=2, column=16).value + 1
    if(clear_num!=1):
        l = 2
        while l <= clear_num:
            for column in sheet.iter_cols(min_row=l, min_col=4, max_row=l, max_col=10):
                for cell in column:
                    cell.value = None
            l += 1
    workbook.save(filename= file_name)

clear('Back-Up API person data.xlsx')