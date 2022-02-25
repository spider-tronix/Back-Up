from cmath import pi
from tf_pose import common
import cv2
from tf_pose.estimator import TfPoseEstimator
from tf_pose.networks import get_graph_path, model_wh
import matplotlib.pyplot as plt
import numpy as np
import math

def anglecalc(keypoint_list,n1,n2,n3):
    
    _,x1,y1 = keypoint_list[n1]
    _,x2,y2 = keypoint_list[n2]
    _,x3,y3 = keypoint_list[n3]

    p1 = np.array([x1,y1]) - np.array([x2,y2])
    p2 = np.array([x3,y3]) - np.array([x2,y2])
    cosine_angle = np.dot(p1, p2) / (np.linalg.norm(p1) * np.linalg.norm(p2))
    angle = np.arccos(cosine_angle)

    return np.degrees(angle)

def distancecalc(keypoint_list):

    _,x1,y1 = keypoint_list[1]
    _,p1,q1 = keypoint_list[8]
    _,p2,q2 = keypoint_list[11]
    x2,y2 = (p1+p2)/2 , (q1+q2)/2

    dist = math.sqrt((x1-y1)**2 + (x2-y2)**2)
    return dist

def slouchdetect():

    #image = 'demo/WhatsApp Image 2022-02-24 at 5.11.42 PM.jpeg'       # image path
    #image = common.read_imgfile(image, None, None)
    cam = cv2.VideoCapture(0)
    res, image = cam.read()
    w, h = model_wh('432x368')       # image resolution - fix at 432x368

    e = TfPoseEstimator(get_graph_path('cmu'), target_size=(w, h))
    humans = e.inference(image, resize_to_default=(w > 0 and h > 0), upsample_size=4.0)
    image = TfPoseEstimator.draw_humans(image, humans, imgcopy=False)
    image = cv2.resize(image,dsize=(432,368))

    people = {}
    for i,human in enumerate(humans):
        coordinates = []
        for j in human.body_parts:
            x = round(human.body_parts[j].x*w,2)
            y = round(human.body_parts[j].y*h,2)
            coordinates.append((j,x,y))
        people['person'+str(i+1)] = coordinates

    people = dict(sorted(people.items(), key=lambda item: item[1][0][1]))
    people_coordinates = {}
    for k in range(0,len(people)):
        people_coordinates['Person'+str(k+1)] = people[str(list(people.items())[k][0])]

    angle = anglecalc(people_coordinates['Person1'],8,1,11)
    dist = distancecalc(people_coordinates['Person1'])

    output = {}
    for i in range(0,len(people_coordinates)):
        output['Person'+str(i+1)] = (anglecalc(people_coordinates['Person'+str(i+1)],8,1,11), distancecalc(people_coordinates['Person'+str(i+1)]))

    plt.imshow(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
    plt.show()

    return output,people_coordinates

op,people_coordinates = slouchdetect()
#print(op)

from openpyxl import load_workbook
import math

# returns square of distance b/w two points
def lengthSquare(X, Y):
        xDiff = X[0] - Y[0]
        yDiff = X[1] - Y[1]
        return xDiff * xDiff + yDiff * yDiff


def Angle(A, B, C):
        # Square of lengths be a2, b2, c2
        a2 = lengthSquare(B, C)
        b2 = lengthSquare(A, C)
        c2 = lengthSquare(A, B)

        # length of sides be a, b, c
        a = math.sqrt(a2);
        b = math.sqrt(b2);
        c = math.sqrt(c2);

        # From Cosine law
        beta = math.acos((a2 + c2 - b2) / (2 * a * c));

        # Converting to degree
        beta = beta * 180 / math.pi;

        # printing all the angles
        return(beta)


workbook = load_workbook(filename = "Back-Up API person data.xlsx")
file_name = "Back-Up API person data.xlsx"
sheet = workbook.active     # Selecting the active sheet

i = 1
n_ppl = len(people_coordinates)

l = 2

while l <= (sheet.cell(row=2, column=16).value+1):
    sheet.cell(row=l, column=12).value = "N" # Resetting presence column
    sheet.cell(row=l, column=14).value = "NA"
    sheet.cell(row=l, column=13).value = "NA"
    l += 1
workbook.save(filename=file_name)

print(len(people_coordinates))
while i <= n_ppl:
    x = "Person"
    y = chr(i+48)
    z = x + y
    Shoulder_gap = people_coordinates[z][5][1] - people_coordinates[z][2][1]
    # print(Shoulder_gap)
    X_left = people_coordinates[z][1][1] - 0.65*Shoulder_gap
    X_right = people_coordinates[z][1][1] + 0.65*Shoulder_gap
    # print(X_left)
    # print(X_right)
    j = 0
    while j < len(people_coordinates[z]):
        if people_coordinates[z][j][0] == 11:
            Person_Coor_11 = people_coordinates[z][j]
        if people_coordinates[z][j][0] == 8:
            Person_Coor_8 = people_coordinates[z][j]
        j += 1

    A = (Person_Coor_11[1], Person_Coor_11[2])  # Left hip
    B = (people_coordinates[z][1][1], people_coordinates[z][1][2])  # Centre of shoulder
    C = (Person_Coor_8[1], Person_Coor_8[2])  # Right hip

    Current_angle = Angle(A, B, C)

    D = ((A[0]+B[0])/2, (A[1]+B[1])/2) # Centre of hip

    Current_length = math.sqrt(lengthSquare(B, D))

    # Allocating person index
    Person_Center = people_coordinates[z][1][1]
    print(Person_Center)
    k = 2
    Person_Index = 0
    while k<=sheet.cell(row=2, column=16).value+1:
        if (Person_Center > sheet.cell(row=k, column=9).value) & (Person_Center < sheet.cell(row=k, column=10).value):
            Person_Index = k-1
            sheet.cell(row=Person_Index + 1, column=12).value = "Y"  # Mention presence
        k += 1
    Refer_length = sheet.cell(row=Person_Index + 1, column=7).value
    Refer_angle = sheet.cell(row=Person_Index + 1, column=8).value
    err_Length_per = sheet.cell(row=2, column=17).value
    err_Angle_per = sheet.cell(row=2, column=18).value
    sheet.cell(row=Person_Index + 1, column=14).value = Current_angle
    sheet.cell(row=Person_Index + 1, column=13).value = Current_length
    if (Current_length < (Refer_length - (Refer_length * err_Length_per/100))) & (Current_angle > (Refer_angle + (Refer_angle * err_Angle_per/100))):
        sheet.cell(row=Person_Index + 1, column=11).value += 1
    else:
        sheet.cell(row=Person_Index + 1, column=11).value = 0
    if sheet.cell(row=Person_Index + 1, column=11).value > 0: # Actually shd be 20 or something
        sheet.cell(row=Person_Index + 1, column=6).value = "Y"
    else:
        sheet.cell(row=Person_Index + 1, column=6).value = "N"

    workbook.save(filename=file_name)
    i += 1
