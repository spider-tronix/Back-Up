from openpyxl import load_workbook
import math
workbook = load_workbook(filename = "Back-Up API person data.xlsx")
file_name = "Back-Up API person data.xlsx"
sheet = workbook.active     # Selecting the active sheet

l = 2
while l <= 5:
    for column in sheet.iter_cols(min_row=l, min_col=4, max_row=l, max_col=14):
        for cell in column:
            cell.value = 0
    l+=1

workbook.save(filename="Back-Up API person data.xlsx")
