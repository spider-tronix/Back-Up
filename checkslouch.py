from calculations import init_coordinates,calc_parameters
from openpyxl import load_workbook

def slouch_detect_write_excel(calibrated_param, output_coordinates, file_name):

    workbook = load_workbook(file_name)
    sheet = workbook.active
    n_ppl = len(output_coordinates)

    l = 2
    while l <= (sheet.cell(row=2, column=16).value + 1):
        sheet.cell(row=l, column=12).value = "N"  # Resetting presence column
        sheet.cell(row=l, column=14).value = "NA"
        sheet.cell(row=l, column=13).value = "NA"
        sheet.cell(row=l, column=4).value = 0
        sheet.cell(row=l, column=5).value = 0
        sheet.cell(row=l, column=11).value = 0
        l += 1
    workbook.save(file_name)
    for i in range(1, n_ppl+1):
        person = 'Person'+str(i)
        coordinates = init_coordinates(output_coordinates,person)
        param = calc_parameters(coordinates)
        k = 2
        Person_Index = 0
        while k <= sheet.cell(row=2, column=16).value + 1:
            if (coordinates['neck'][0] > sheet.cell(row=k, column=9).value) & (coordinates['neck'][0] < sheet.cell(row=k, column=10).value) & (coordinates['neck'][1] < sheet.cell(row=k, column=19).value) & (coordinates['neck'][1] > sheet.cell(row=k, column=20).value):
                Person_Index = k - 1
                sheet.cell(row=Person_Index + 1, column=12).value = "Y" 
            k += 1

        if Person_Index != 0:
            # Straight facing
            if sheet.cell(row=Person_Index + 1, column=21).value == 0:
                Refer_length = sheet.cell(row=Person_Index + 1, column=7).value
                Refer_angle = sheet.cell(row=Person_Index + 1, column=8).value
                err_Length_per = sheet.cell(row=2, column=17).value
                err_Angle_per = sheet.cell(row=2, column=18).value
                sheet.cell(row=Person_Index + 1, column=23).value = param['nose_to_neck_length']
                sheet.cell(row=Person_Index + 1, column=14).value = param['torso_angle']
                sheet.cell(row=Person_Index + 1, column=13).value = param['spine_length']
                Length_slouch = param['spine_length'] < (Refer_length - (Refer_length * err_Length_per / 100))
                Angle_slouch = param['torso_angle'] > (Refer_angle + (Refer_angle * err_Angle_per / 100))
                Neck_slouch = (sheet.cell(row= Person_Index + 1, column=22).value - param['nose_to_neck_length']) > ((5/100)*sheet.cell(row=Person_Index + 1, column=22).value)
                if (Length_slouch & Angle_slouch) or Neck_slouch:
                    sheet.cell(row=Person_Index + 1, column=11).value += 1
                else:
                    sheet.cell(row=Person_Index + 1, column=11).value = 0
                if sheet.cell(row=Person_Index + 1, column=11).value > 0:  # Actually shd be 20 or something (Continuous slouching)
                    sheet.cell(row=Person_Index + 1, column=6).value = "Y"
                else:
                    sheet.cell(row=Person_Index + 1, column=6).value = "N"

            # Side facing
            elif(sheet.cell(row=Person_Index + 1, column=21).value == 1):
                sheet.cell(row=Person_Index + 1, column=25).value = param['neck_angle']
                sheet.cell(row=Person_Index + 1, column=27).value = param['back_angle']
                if param['neck_angle']>sheet.cell(row=Person_Index + 1, column=24).value or param['back_angle']>sheet.cell(row=Person_Index + 1, column=26).value:
                    sheet.cell(row=Person_Index + 1, column=11).value += 1
                else:
                    sheet.cell(row=Person_Index + 1, column=11).value = 0
                if sheet.cell(row=Person_Index + 1, column=11).value > 0:
                    sheet.cell(row=Person_Index + 1, column=6).value = "Y"
                    sheet.cell(row=Person_Index + 1, column=11).value = 0
                else:
                    sheet.cell(row=Person_Index + 1, column=6).value = "N"

        workbook.save(file_name)
