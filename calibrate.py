from calculations import init_coordinates,calc_parameters
from openpyxl import load_workbook

def calibrate_write_excel(output_coordinates,file_name):
    
    n_ppl = len(output_coordinates)
    workbook = load_workbook(filename= file_name)
    sheet = workbook.active
    calibrated_param = {}

    for i in range(1,n_ppl+1):
        person = 'Person'+str(i)
        coordinates = init_coordinates(output_coordinates, person)
        param = calc_parameters(coordinates)
        calibrated_param[person] = calc_parameters(coordinates)

        # Check for sideways slouching
        sheet.cell(row=i + 1, column=21).value = 0
        if coordinates['nose'][0] < min(coordinates['right_shoulder'][0],coordinates['left_shoulder'][0]) or coordinates['nose'][0] > max(coordinates['right_shoulder'][0],coordinates['left_shoulder'][0]):
            sheet.cell(row=i + 1, column=21).value = 1
            sheet.cell(row=i + 1, column=24).value = param['neck_angle']
            sheet.cell(row=i + 1, column=26).value = param['back_angle']
        
        sheet.cell(row=i + 1, column=22).value = param['nose_to_neck_length']
        sheet.cell(row=i + 1, column=9).value = param['X_left']
        sheet.cell(row=i + 1, column=10).value = param['X_right']
        sheet.cell(row=i + 1, column=8).value = param['torso_angle']
        sheet.cell(row=i + 1, column=7).value = param['spine_length']
        sheet.cell(row=i + 1, column=19).value = param['Y_down']
        sheet.cell(row=i + 1, column=20).value = param['Y_up']
        sheet.cell(row=2, column=16).value = n_ppl

    workbook.save(filename=file_name)
    return calibrated_param